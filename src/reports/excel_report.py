"""Excel report generator — multi-sheet workbook with charts.

Uses openpyxl for .xlsx generation. Creates:
- Summary sheet (key metrics per model)
- Per-model sheets (detailed trades, equity curve data)
- Comparison sheet (side-by-side metrics)
- Charts sheet (embedded chart images from data)
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import datetime, timezone

from src.core.logging import get_logger

log = get_logger(__name__)


@dataclass
class PortfolioReportData:
    """Data for one portfolio in the report."""
    model: str
    architecture: str
    market: str
    initial_capital: float
    final_value: float
    total_return_pct: float
    sharpe_ratio: float
    max_drawdown: float
    win_rate: float
    total_trades: int
    total_api_cost: float
    cost_adjusted_alpha: float
    equity_curve: list[float] = field(default_factory=list)
    trade_log: list[dict[str, object]] = field(default_factory=list)


@dataclass
class ReportConfig:
    """Configuration for report generation."""
    title: str = "ATLAS Benchmark Report"
    author: str = "ATLAS System"
    generated_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    include_trade_log: bool = True
    include_equity_curves: bool = True
    include_charts: bool = True


class ExcelReportGenerator:
    """Generate comprehensive Excel reports."""

    def __init__(self, config: ReportConfig | None = None) -> None:
        self._config = config or ReportConfig()

    def generate(self, portfolios: list[PortfolioReportData]) -> bytes:
        """Generate Excel workbook as bytes.

        Returns bytes that can be written to file or sent as HTTP response.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import LineChart, BarChart, Reference
        except ImportError:
            log.warning("openpyxl_not_installed", msg="Using fallback CSV-style output")
            return self._generate_csv_fallback(portfolios)

        wb = Workbook()

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary_sheet(ws_summary, portfolios, Font, PatternFill, Alignment, Border, Side, get_column_letter)

        # Sheet 2: Comparison
        ws_compare = wb.create_sheet("Model Comparison")
        self._write_comparison_sheet(ws_compare, portfolios, Font, PatternFill, Alignment)

        # Sheet 3: Equity Curves (data for charting)
        if self._config.include_equity_curves:
            ws_equity = wb.create_sheet("Equity Curves")
            self._write_equity_sheet(ws_equity, portfolios, LineChart, Reference, Font, PatternFill)

        # Sheet 4: Trade Log
        if self._config.include_trade_log:
            ws_trades = wb.create_sheet("Trade Log")
            self._write_trade_log_sheet(ws_trades, portfolios, Font, PatternFill)

        # Sheet 5: Charts
        if self._config.include_charts:
            ws_charts = wb.create_sheet("Charts")
            self._write_charts_sheet(ws_charts, portfolios, wb, BarChart, LineChart, Reference)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        log.info("excel_report_generated", sheets=len(wb.sheetnames), portfolios=len(portfolios))
        return buf.getvalue()

    def _write_summary_sheet(self, ws: object, portfolios: list[PortfolioReportData],
                             Font: type, PatternFill: type, Alignment: type,
                             Border: type, Side: type, get_column_letter: object) -> None:
        """Write the summary overview sheet."""
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = self._config.title
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated: {self._config.generated_at.strftime('%Y-%m-%d %H:%M UTC')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws['A3'] = f"Author: {self._config.author}"

        # Headers
        headers = ["Model", "Architecture", "Market", "Initial Capital", "Final Value",
                   "Return %", "Sharpe", "Max DD %", "Win Rate %", "Trades", "API Cost", "CAA"]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, p in enumerate(portfolios, 6):
            ws.cell(row=row_idx, column=1, value=p.model)
            ws.cell(row=row_idx, column=2, value=p.architecture)
            ws.cell(row=row_idx, column=3, value=p.market)
            ws.cell(row=row_idx, column=4, value=round(p.initial_capital, 2))
            ws.cell(row=row_idx, column=5, value=round(p.final_value, 2))
            ws.cell(row=row_idx, column=6, value=round(p.total_return_pct, 2))
            ws.cell(row=row_idx, column=7, value=round(p.sharpe_ratio, 4))
            ws.cell(row=row_idx, column=8, value=round(p.max_drawdown * 100, 2))
            ws.cell(row=row_idx, column=9, value=round(p.win_rate * 100, 1))
            ws.cell(row=row_idx, column=10, value=p.total_trades)
            ws.cell(row=row_idx, column=11, value=round(p.total_api_cost, 4))
            ws.cell(row=row_idx, column=12, value=round(p.cost_adjusted_alpha, 4))

            # Alternate row colors
            if row_idx % 2 == 0:
                for col in range(1, 13):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

        # Auto-width columns
        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_comparison_sheet(self, ws: object, portfolios: list[PortfolioReportData],
                                Font: type, PatternFill: type, Alignment: type) -> None:
        """Write model comparison sheet."""
        ws['A1'] = "Model Comparison"
        ws['A1'].font = Font(size=14, bold=True)

        # Group by model
        models: dict[str, list[PortfolioReportData]] = {}
        for p in portfolios:
            models.setdefault(p.model, []).append(p)

        row = 3
        for model, plist in models.items():
            ws.cell(row=row, column=1, value=model).font = Font(bold=True, size=12)
            row += 1
            headers = ["Architecture", "Market", "Return %", "Sharpe", "Max DD %", "CAA"]
            for ci, h in enumerate(headers, 1):
                ws.cell(row=row, column=ci, value=h).font = Font(bold=True)
            row += 1
            for p in plist:
                ws.cell(row=row, column=1, value=p.architecture)
                ws.cell(row=row, column=2, value=p.market)
                ws.cell(row=row, column=3, value=round(p.total_return_pct, 2))
                ws.cell(row=row, column=4, value=round(p.sharpe_ratio, 4))
                ws.cell(row=row, column=5, value=round(p.max_drawdown * 100, 2))
                ws.cell(row=row, column=6, value=round(p.cost_adjusted_alpha, 4))
                row += 1
            row += 1

    def _write_equity_sheet(self, ws: object, portfolios: list[PortfolioReportData],
                            LineChart: type, Reference: type, Font: type,
                            PatternFill: type) -> None:
        """Write equity curve data + chart."""
        ws['A1'] = "Equity Curves"
        ws['A1'].font = Font(size=14, bold=True)

        # Headers: Period, Model1, Model2, ...
        ws.cell(row=3, column=1, value="Period")
        for ci, p in enumerate(portfolios, 2):
            ws.cell(row=3, column=ci, value=f"{p.model}/{p.architecture}")

        # Data
        max_len = max((len(p.equity_curve) for p in portfolios), default=0)
        for ri in range(max_len):
            ws.cell(row=ri + 4, column=1, value=ri + 1)
            for ci, p in enumerate(portfolios, 2):
                if ri < len(p.equity_curve):
                    ws.cell(row=ri + 4, column=ci, value=round(p.equity_curve[ri], 2))

        # Create line chart if data exists
        if max_len > 0 and len(portfolios) > 0:
            chart = LineChart()
            chart.title = "Equity Curves"
            chart.y_axis.title = "Portfolio Value"
            chart.x_axis.title = "Period"
            chart.width = 30
            chart.height = 15

            for ci in range(2, len(portfolios) + 2):
                values = Reference(ws, min_col=ci, min_row=3, max_row=max_len + 3)
                chart.add_data(values, titles_from_data=True)

            ws.add_chart(chart, "A" + str(max_len + 6))

    def _write_trade_log_sheet(self, ws: object, portfolios: list[PortfolioReportData],
                               Font: type, PatternFill: type) -> None:
        """Write combined trade log."""
        ws['A1'] = "Trade Log"
        ws['A1'].font = Font(size=14, bold=True)

        headers = ["Model", "Architecture", "Timestamp", "Symbol", "Action",
                   "Quantity", "Price", "Commission", "Realized PnL"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=3, column=ci, value=h).font = Font(bold=True)

        row = 4
        for p in portfolios:
            for trade in p.trade_log:
                ws.cell(row=row, column=1, value=p.model)
                ws.cell(row=row, column=2, value=p.architecture)
                ws.cell(row=row, column=3, value=str(trade.get("timestamp", "")))
                ws.cell(row=row, column=4, value=str(trade.get("symbol", "")))
                ws.cell(row=row, column=5, value=str(trade.get("action", "")))
                ws.cell(row=row, column=6, value=float(trade.get("quantity", 0)))
                ws.cell(row=row, column=7, value=float(trade.get("price", 0)))
                ws.cell(row=row, column=8, value=float(trade.get("commission", 0)))
                ws.cell(row=row, column=9, value=float(trade.get("realized_pnl", 0)))
                row += 1

    def _write_charts_sheet(self, ws: object, portfolios: list[PortfolioReportData],
                            wb: object, BarChart: type, LineChart: type,
                            Reference: type) -> None:
        """Write charts (return comparison bar chart)."""
        ws['A1'] = "Performance Charts"

        # Return comparison data
        ws.cell(row=3, column=1, value="Model")
        ws.cell(row=3, column=2, value="Return %")
        ws.cell(row=3, column=3, value="Sharpe")
        ws.cell(row=3, column=4, value="CAA")

        for ri, p in enumerate(portfolios, 4):
            ws.cell(row=ri, column=1, value=f"{p.model}/{p.architecture}")
            ws.cell(row=ri, column=2, value=round(p.total_return_pct, 2))
            ws.cell(row=ri, column=3, value=round(p.sharpe_ratio, 4))
            ws.cell(row=ri, column=4, value=round(p.cost_adjusted_alpha, 4))

        if portfolios:
            chart = BarChart()
            chart.title = "Return Comparison"
            chart.y_axis.title = "Return %"
            chart.width = 25
            chart.height = 12

            data = Reference(ws, min_col=2, min_row=3, max_row=len(portfolios) + 3)
            cats = Reference(ws, min_col=1, min_row=4, max_row=len(portfolios) + 3)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "A" + str(len(portfolios) + 6))

    def _generate_csv_fallback(self, portfolios: list[PortfolioReportData]) -> bytes:
        """CSV fallback when openpyxl is not installed."""
        lines = ["Model,Architecture,Market,Return%,Sharpe,MaxDD%,WinRate%,Trades,APICost,CAA"]
        for p in portfolios:
            lines.append(f"{p.model},{p.architecture},{p.market},{p.total_return_pct:.2f},"
                        f"{p.sharpe_ratio:.4f},{p.max_drawdown*100:.2f},{p.win_rate*100:.1f},"
                        f"{p.total_trades},{p.total_api_cost:.4f},{p.cost_adjusted_alpha:.4f}")
        return "\n".join(lines).encode("utf-8")
